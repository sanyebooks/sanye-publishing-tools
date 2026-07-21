#!/usr/bin/env python3
"""Validate title length and cross-sheet title reuse in a children's book workbook."""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

from openpyxl import load_workbook


MAIN_SHEET = "卖点表"
CONTENT_SHEET = "小红书笔记&图像提示词"
TITLE_HEADER = "网感化后的标题"
INDEX_HEADER = "序号"
AGE_RANGE_PATTERN = re.compile(r"\d+\s*(?:[-—–~～]|至|到)\s*\d+\s*岁")


def visible_length(value: object) -> int:
    text = "" if value is None else str(value)
    return sum(1 for char in text if char not in "\r\n")


def find_header(ws, header: str) -> int:
    for cell in ws[1]:
        if cell.value == header:
            return cell.column
    raise ValueError(f"{ws.title}: 缺少列 {header!r}")


def validate(path: Path, max_length: int) -> list[str]:
    wb = load_workbook(path, data_only=False)
    errors: list[str] = []

    if MAIN_SHEET not in wb.sheetnames:
        return [f"缺少工作表 {MAIN_SHEET!r}"]

    main = wb[MAIN_SHEET]
    try:
        main_title_col = find_header(main, TITLE_HEADER)
    except ValueError as exc:
        return [str(exc)]

    main_titles: list[str] = []
    for row in range(2, main.max_row + 1):
        raw = main.cell(row, main_title_col).value
        title = "" if raw is None else str(raw).strip()
        main_titles.append(title)
        if not title:
            errors.append(f"{MAIN_SHEET}!R{row}: 标题为空")
            continue
        if "\n" in title or "\r" in title:
            errors.append(f"{MAIN_SHEET}!R{row}: 标题含换行")
        if AGE_RANGE_PATTERN.search(title):
            errors.append(
                f"{MAIN_SHEET}!R{row}: 标题使用宽泛年龄范围，应改为单点年龄与具体孩子画像: {title}"
            )
        length = visible_length(title)
        if length > max_length:
            errors.append(
                f"{MAIN_SHEET}!R{row}: {length}个可见字符，超过上限{max_length}: {title}"
            )

    if CONTENT_SHEET not in wb.sheetnames:
        return errors

    content = wb[CONTENT_SHEET]
    try:
        content_title_col = find_header(content, TITLE_HEADER)
        index_col = find_header(content, INDEX_HEADER)
    except ValueError as exc:
        errors.append(str(exc))
        return errors

    for row in range(2, content.max_row + 1):
        raw_index = content.cell(row, index_col).value
        raw_title = content.cell(row, content_title_col).value
        title = "" if raw_title is None else str(raw_title).strip()
        if raw_index in (None, "") and not title:
            continue
        try:
            index = int(raw_index)
        except (TypeError, ValueError):
            errors.append(f"{CONTENT_SHEET}!R{row}: 序号无效 {raw_index!r}")
            continue
        if index < 1 or index > len(main_titles):
            errors.append(f"{CONTENT_SHEET}!R{row}: 序号{index}超出主表范围")
            continue
        expected = main_titles[index - 1]
        if title != expected:
            errors.append(
                f"{CONTENT_SHEET}!R{row}: 标题与主表第{index}条不一致; "
                f"当前={title!r}, 应为={expected!r}"
            )

    return errors


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--max-length", type=int, default=20)
    args = parser.parse_args()

    if not args.workbook.is_file():
        print(f"ERROR: 文件不存在: {args.workbook}", file=sys.stderr)
        return 2

    try:
        errors = validate(args.workbook, args.max_length)
    except Exception as exc:  # keep CLI failure readable
        print(f"ERROR: 无法检查工作簿: {exc}", file=sys.stderr)
        return 2

    if errors:
        print(f"FAIL: 发现{len(errors)}个标题问题")
        for error in errors:
            print(f"- {error}")
        return 1

    print(
        f"PASS: 主表标题全部≤{args.max_length}个可见字符且无宽泛年龄范围，"
        "Sheet3标题与主表一致"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
